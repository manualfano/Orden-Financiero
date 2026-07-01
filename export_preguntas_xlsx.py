import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

data = [
    {"eslabon":"Costos y Precios","pregunta":"¿Sabés cuánto te cuesta realmente producir o dar cada servicio?",
     "A":"Sí, lo tengo calculado y lo actualizo seguido","B":"A veces, pero no siempre lo tengo actualizado","C":"No, lo calculo a ojo",
     "ok":"Perfecto. El costeo actualizado es la base de todo lo demás.",
     "warn":"Vas por buen camino, pero sin un número exacto trabajás a intuición. Ese es el primer agujero.",
     "crit":"Es lo más común en PyMEs. Sin ese número, cualquier precio que ponés es una apuesta."},
    {"eslabon":"Costos y Precios","pregunta":"¿Tu precio de venta está calculado sobre tus costos reales o lo fijás por referencia al mercado?",
     "A":"Sí, lo tengo claro y lo hago sistemáticamente","B":"A veces, pero no de manera consistente","C":"No, no lo tengo resuelto",
     "ok":"Muy bien. Esto es una fortaleza de tu negocio.",
     "warn":"Está en proceso, pero la inconsistencia genera brechas. Vale la pena sistematizarlo.",
     "crit":"Este es un punto crítico. Sin resolverlo, el resto del negocio se ve afectado."},
    {"eslabon":"Costos y Precios","pregunta":"¿Revisás y actualizás tus precios cuando suben tus costos?",
     "A":"Sí, lo tengo claro y lo hago sistemáticamente","B":"A veces, pero no de manera consistente","C":"No, no lo tengo resuelto",
     "ok":"Muy bien. Esto es una fortaleza de tu negocio.",
     "warn":"Está en proceso, pero la inconsistencia genera brechas. Vale la pena sistematizarlo.",
     "crit":"Este es un punto crítico. Sin resolverlo, el resto del negocio se ve afectado."},
    {"eslabon":"Resultado Económico","pregunta":"¿Sabés exactamente cuánto ganó tu negocio el mes pasado?",
     "A":"Sí, lo tengo claro y lo hago sistemáticamente","B":"A veces, pero no de manera consistente","C":"No, no lo tengo resuelto",
     "ok":"Muy bien. Esto es una fortaleza de tu negocio.",
     "warn":"Está en proceso, pero la inconsistencia genera brechas. Vale la pena sistematizarlo.",
     "crit":"Este es un punto crítico. Sin resolverlo, el resto del negocio se ve afectado."},
    {"eslabon":"Resultado Económico","pregunta":"¿Separás las finanzas del negocio de tus gastos personales?",
     "A":"Sí, lo tengo claro y lo hago sistemáticamente","B":"A veces, pero no de manera consistente","C":"No, no lo tengo resuelto",
     "ok":"Muy bien. Esto es una fortaleza de tu negocio.",
     "warn":"Está en proceso, pero la inconsistencia genera brechas. Vale la pena sistematizarlo.",
     "crit":"Este es un punto crítico. Sin resolverlo, el resto del negocio se ve afectado."},
    {"eslabon":"Resultado Económico","pregunta":"¿Tenés un cierre mensual que te muestre ventas, costos y ganancia neta?",
     "A":"Sí, lo tengo claro y lo hago sistemáticamente","B":"A veces, pero no de manera consistente","C":"No, no lo tengo resuelto",
     "ok":"Muy bien. Esto es una fortaleza de tu negocio.",
     "warn":"Está en proceso, pero la inconsistencia genera brechas. Vale la pena sistematizarlo.",
     "crit":"Este es un punto crítico. Sin resolverlo, el resto del negocio se ve afectado."},
    {"eslabon":"Cash Flow","pregunta":"¿Llegás a fin de mes con plata disponible para pagar sueldos y proveedores sin apurarte?",
     "A":"Sí, lo tengo claro y lo hago sistemáticamente","B":"A veces, pero no de manera consistente","C":"No, no lo tengo resuelto",
     "ok":"Muy bien. Esto es una fortaleza de tu negocio.",
     "warn":"Está en proceso, pero la inconsistencia genera brechas. Vale la pena sistematizarlo.",
     "crit":"Este es un punto crítico. Sin resolverlo, el resto del negocio se ve afectado."},
    {"eslabon":"Cash Flow","pregunta":"¿Sabés con anticipación cuándo van a ser tus próximos gastos fuertes?",
     "A":"Sí, lo tengo claro y lo hago sistemáticamente","B":"A veces, pero no de manera consistente","C":"No, no lo tengo resuelto",
     "ok":"Muy bien. Esto es una fortaleza de tu negocio.",
     "warn":"Está en proceso, pero la inconsistencia genera brechas. Vale la pena sistematizarlo.",
     "crit":"Este es un punto crítico. Sin resolverlo, el resto del negocio se ve afectado."},
    {"eslabon":"Cash Flow","pregunta":"¿Tus cobros y pagos están equilibrados o siempre pagás antes de cobrar?",
     "A":"Sí, lo tengo claro y lo hago sistemáticamente","B":"A veces, pero no de manera consistente","C":"No, no lo tengo resuelto",
     "ok":"Muy bien. Esto es una fortaleza de tu negocio.",
     "warn":"Está en proceso, pero la inconsistencia genera brechas. Vale la pena sistematizarlo.",
     "crit":"Este es un punto crítico. Sin resolverlo, el resto del negocio se ve afectado."},
    {"eslabon":"Indicadores de Gestión","pregunta":"¿Revisás los números de tu negocio al menos una vez por mes?",
     "A":"Sí, lo tengo claro y lo hago sistemáticamente","B":"A veces, pero no de manera consistente","C":"No, no lo tengo resuelto",
     "ok":"Muy bien. Esto es una fortaleza de tu negocio.",
     "warn":"Está en proceso, pero la inconsistencia genera brechas. Vale la pena sistematizarlo.",
     "crit":"Este es un punto crítico. Sin resolverlo, el resto del negocio se ve afectado."},
    {"eslabon":"Indicadores de Gestión","pregunta":"¿Tenés algún indicador (ventas, margen, CMV) que seguís regularmente?",
     "A":"Sí, lo tengo claro y lo hago sistemáticamente","B":"A veces, pero no de manera consistente","C":"No, no lo tengo resuelto",
     "ok":"Muy bien. Esto es una fortaleza de tu negocio.",
     "warn":"Está en proceso, pero la inconsistencia genera brechas. Vale la pena sistematizarlo.",
     "crit":"Este es un punto crítico. Sin resolverlo, el resto del negocio se ve afectado."},
    {"eslabon":"Indicadores de Gestión","pregunta":"¿Tomás decisiones de inversión o contratación basándote en datos concretos?",
     "A":"Sí, lo tengo claro y lo hago sistemáticamente","B":"A veces, pero no de manera consistente","C":"No, no lo tengo resuelto",
     "ok":"Muy bien. Esto es una fortaleza de tu negocio.",
     "warn":"Está en proceso, pero la inconsistencia genera brechas. Vale la pena sistematizarlo.",
     "crit":"Este es un punto crítico. Sin resolverlo, el resto del negocio se ve afectado."},
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Preguntas Diagnóstico"

headers = ["#","Eslabón","Pregunta","Opción A (3 pts)","Opción B (2 pts)","Opción C (1 pt)",
           "Reacción A (ok)","Reacción B (warn)","Reacción C (crit)"]
ws.append(headers)

header_fill = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = header_fill
    c.alignment = Alignment(vertical="center", wrap_text=True)

for i, q in enumerate(data, 1):
    ws.append([i, q["eslabon"], q["pregunta"], q["A"], q["B"], q["C"], q["ok"], q["warn"], q["crit"]])

widths = [4, 20, 38, 30, 30, 26, 34, 38, 36]
for col, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = w

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

ws.freeze_panes = "A2"

out_path = r"C:\Users\usuario\Desktop\orden-financiero\preguntas_diagnostico.xlsx"
wb.save(out_path)
print("Guardado en:", out_path)
